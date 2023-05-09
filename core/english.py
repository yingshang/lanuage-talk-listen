from openpyxl import load_workbook
from core.models import insert_toefl_independent
from core.scenes import school_scenes,academic_scenes

def toefl_load_independent_file(file):
    # 加载Excel文件
    workbook = load_workbook(file)

    # 选择第一个工作表
    worksheet = workbook.worksheets[0]
    num = 0
    total = 0
    # 读取第一列（A列）从第二行开始的数据
    for row in worksheet.iter_rows(min_row=2, min_col=1, values_only=True):
        cell_value = row[0]
        n = insert_toefl_independent(cell_value)
        num = num+n
        total = total+1
    return total,num



def get_listen_school_scene():
    content = []
    for key,value in school_scenes.items():
        sc = [
            [
            {
                "tag": "text",
                "text": f"{key} ",
                "style": ["bold", "underline"]
            },
            {
                "tag": "text",
                "text": ";".join(value),
                "style": []
            }
        ],
            [
                {
                    "tag": "text",
                    "text": "",
                    "style": []
                }
            ]
        ]
        content = content+sc

    resp_text = {
        "title": "学校听力场景总览",
        "content": content
    }
    return resp_text

def get_listen_academic_scene():
    content = []
    for ac in academic_scenes:

        acad = [
            [

                {
                    "tag": "text",
                    "text": ac,
                    "style": []
                }
            ],
            [
                {
                    "tag": "text",
                    "text": "",
                    "style": []
                }
            ]
        ]
        content = content + acad



    resp_text = {
        "title": "讲座听力场景总览",
        "content": content
    }
    return resp_text